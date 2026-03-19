import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import copy, io, zipfile, os, re, uuid, traceback
from pathlib import Path

# ══════════════════════════════════════════════════════════════════
# 매핑
# ══════════════════════════════════════════════════════════════════
COMPETENCY_MAP = {
    "Position":     [9, 10, 17, 18, 22, 31],
    "Personality":  [5, 13, 14, 21, 24, 28],
    "Relationship": [6,  7, 23, 25, 30, 32],
    "Results":      [8, 16, 29, 33],
    "Development":  [4, 12, 20, 27],
    "Principles":   [11, 15, 19, 26],
}
SKILL_MAP = {
    "우호성":     [4, 12, 20, 27],
    "동기유발":   [5, 13, 21, 28],
    "자문":       [6, 14],
    "협력제휴":   [7, 15, 22, 29],
    "협상거래":   [8, 16, 23, 30],
    "합리적설득": [9, 17, 24, 31],
    "합법화":     [10, 18, 25, 32],
    "강요":       [11, 19, 26, 33],
}
SOFT_SKILLS = ["우호성", "동기유발", "자문"]
HARD_SKILLS = ["협력제휴", "협상거래", "합리적설득", "합법화", "강요"]
COMP_ROW  = {"Position":4,"Personality":5,"Relationship":6,
              "Results":7,"Development":8,"Principles":9}
SKILL_ROW = {"우호성":12,"동기유발":13,"자문":14,"협력제휴":15,
              "협상거래":16,"합리적설득":17,"합법화":18,"강요":19}

# ══════════════════════════════════════════════════════════════════
# 계산
# ══════════════════════════════════════════════════════════════════
def avg_rows(scores, rows):
    vals = [float(scores.get(str(r-3), 0)) for r in rows]
    return round(sum(vals)/len(vals), 2) if vals else 0.0

def compute(scores):
    c = {k: avg_rows(scores, v) for k, v in COMPETENCY_MAP.items()}
    s = {k: avg_rows(scores, v) for k, v in SKILL_MAP.items()}
    return {"competency": c, "skill_raw": s,
            "soft_avg": round(sum(s[k] for k in SOFT_SKILLS)/3, 2),
            "hard_avg": round(sum(s[k] for k in HARD_SKILLS)/5, 2)}

# ══════════════════════════════════════════════════════════════════
# 파싱
# ══════════════════════════════════════════════════════════════════
def parse_people(raw: bytes) -> list:
    df = pd.read_excel(io.BytesIO(raw), header=0)
    out = []
    for _, row in df.iterrows():
        name = str(row.iloc[1]).strip()
        if not name or name.lower() == "nan":
            continue
        scores = {str(q): float(row.iloc[q+1]) if pd.notna(row.iloc[q+1]) else 0.0
                  for q in range(1, 31)}
        out.append({"name": name, "scores": scores})
    return out

# ══════════════════════════════════════════════════════════════════
# 엑셀 생성
# ══════════════════════════════════════════════════════════════════
def _copy_ws(wb, src, title):
    ws = wb.create_sheet(title=title)
    for cl, cd in src.column_dimensions.items():
        ws.column_dimensions[cl].width = cd.width
    for rn, rd in src.row_dimensions.items():
        ws.row_dimensions[rn].height = rd.height
    for row in src.iter_rows():
        for cell in row:
            nc = ws.cell(row=cell.row, column=cell.column)
            nc.value = cell.value
            if cell.has_style:
                nc.font=copy.copy(cell.font); nc.border=copy.copy(cell.border)
                nc.fill=copy.copy(cell.fill); nc.number_format=cell.number_format
                nc.protection=copy.copy(cell.protection); nc.alignment=copy.copy(cell.alignment)
    for m in src.merged_cells.ranges:
        ws.merge_cells(str(m))
    return ws

def build_excel(people, excel_tpl: bytes) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for p in people:
        src = load_workbook(io.BytesIO(excel_tpl)).worksheets[0]
        ws  = _copy_ws(wb, src, p["name"][:31])
        r   = compute(p["scores"])
        ws.cell(1,1).value = p["name"]
        for q in range(1, 31):
            ws.cell(q+3, 3).value = float(p["scores"].get(str(q), 0))
        for k, rl in COMPETENCY_MAP.items():
            avg = r["competency"][k]
            ws.cell(COMP_ROW[k],7).value = round(avg*len(rl),2); ws.cell(COMP_ROW[k],7).number_format="0.00"
            ws.cell(COMP_ROW[k],8).value = avg;                  ws.cell(COMP_ROW[k],8).number_format="0.00"
        for k, rl in SKILL_MAP.items():
            avg = r["skill_raw"][k]
            ws.cell(SKILL_ROW[k],7).value = round(avg*len(rl),2); ws.cell(SKILL_ROW[k],7).number_format="0.00"
            ws.cell(SKILL_ROW[k],8).value = avg;                   ws.cell(SKILL_ROW[k],8).number_format="0.00"
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
# PPT 생성
# 베이스: 2슬라이드짜리 template_pptx.pptx
# slide1 → 1번째 사람, slide2 → 2번째 사람
# 3번째 이후 → slide2 구조 복제해서 추가
# ══════════════════════════════════════════════════════════════════
def _inject_table(slide_str, shape_name, rows_data):
    start = slide_str.find(f'name="{shape_name}"')
    if start == -1: return slide_str
    end = slide_str.find('</p:graphicFrame>', start) + len('</p:graphicFrame>')
    section = slide_str[start:end]
    row_matches = list(re.finditer(r'<a:tr\b[^>]*>.*?</a:tr>', section, re.DOTALL))
    new_section = section
    for i, (label, val) in enumerate(rows_data):
        ri = i+1
        if ri >= len(row_matches): break
        orig_row = row_matches[ri].group(0)
        cells = list(re.finditer(r'<a:tc>.*?</a:tc>', orig_row, re.DOTALL))
        if len(cells) < 2: continue
        new_row = orig_row
        for ci, text in enumerate([str(label), f"{float(val):.2f}" if isinstance(val, float) else str(val)]):
            if ci >= len(cells): break
            cell_xml = cells[ci].group(0)
            p_m = re.search(r'<a:p>.*?</a:p>', cell_xml, re.DOTALL)
            if p_m:
                new_p = f'<a:p><a:r><a:rPr lang="ko-KR" sz="900" dirty="0"/><a:t>{text}</a:t></a:r></a:p>'
                new_row = new_row.replace(cell_xml, cell_xml.replace(p_m.group(0), new_p, 1), 1)
                cells = list(re.finditer(r'<a:tc>.*?</a:tc>', new_row, re.DOTALL))
        new_section = new_section.replace(orig_row, new_row, 1)
    return slide_str[:start] + new_section + slide_str[end:]

def _replace_chart_vals(chart_bytes, new_vals):
    s = chart_bytes.decode('utf-8')
    val_m = re.search(r'(<c:val>.*?<c:numCache>)(.*?)(</c:numCache>.*?</c:val>)', s, re.DOTALL)
    if not val_m: return chart_bytes
    before = re.sub(r'<c:ptCount val="\d+"/>', f'<c:ptCount val="{len(new_vals)}"/>', val_m.group(1))
    fmt = re.search(r'<c:formatCode>[^<]*</c:formatCode>', val_m.group(2))
    fmt_tag = fmt.group(0) if fmt else '<c:formatCode>0.00</c:formatCode>'
    pts = ''.join(f'<c:pt idx="{i}"><c:v>{v}</c:v></c:pt>' for i,v in enumerate(new_vals))
    return (s[:val_m.start()] + before + f'{fmt_tag}<c:ptCount val="{len(new_vals)}"/>{pts}' + val_m.group(3) + s[val_m.end():]).encode('utf-8')

def _new_guids(s):
    for g in set(re.findall(r'\{[0-9A-F]{8}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{4}-[0-9A-F]{12}\}', s)):
        s = s.replace(g, '{'+str(uuid.uuid4()).upper()+'}')
    return s

def _ws_name(n): return f"Microsoft_Excel_Worksheet{n if n>0 else ''}.xlsx"

# ── 차트 색상 / 동그라미 위치 상수 (역산으로 구한 실제 플롯 영역) ──
_PHASE_PLOT_X = 2412488;  _PHASE_BAR_W = 1244294   # 6개 막대
_STRAT_PLOT_X = 2351917;  _STRAT_BAR_W = 763098    # 10개 막대
_CIRCLE_PHASE_Y  = 3213000;  _CIRCLE_PHASE_CY  = 437638
_CIRCLE_STRAT_Y  = 6021000;  _CIRCLE_STRAT_CY  = 465643

def _bar_cx_phase(idx):
    return int(_PHASE_PLOT_X + (idx + 0.5) * _PHASE_BAR_W)

def _bar_cx_strat(idx):
    return int(_STRAT_PLOT_X + (idx + 0.5) * _STRAT_BAR_W)

def _update_chart_phase_colors(chart_bytes, vals):
    """최대→파란색(4480B1), 최소→빨간색(C00000). 동값이면 모두 색칠."""
    s = chart_bytes.decode('utf-8')
    max_val = max(vals); min_val = min(vals)
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    dpts = ''
    # 최대값 인덱스 모두 파란색 (최대=최소인 경우 skip)
    if max_val != min_val:
        for idx in [i for i,v in enumerate(vals) if v == max_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="4480B1"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
        for idx in [i for i,v in enumerate(vals) if v == min_val]:
            dpts += (f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                     f'<c:spPr><a:solidFill><a:srgbClr val="C00000"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>')
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _update_chart_strategy_colors(chart_bytes, vals):
    """전체 노란색(FFD000), 소프트평균(idx=3)·하드평균(idx=9)만 남색(2D5576)"""
    s = chart_bytes.decode('utf-8')
    s = re.sub(r'<c:dPt>.*?</c:dPt>', '', s, flags=re.DOTALL)
    s = re.sub(
        r'(<c:spPr>)<a:solidFill>.*?</a:solidFill>',
        r'\1<a:solidFill><a:srgbClr val="FFD000"/></a:solidFill>',
        s, count=1, flags=re.DOTALL
    )
    dpts = ''
    for idx in [3, 9]:
        if idx < len(vals):
            dpts += (
                f'<c:dPt><c:idx val="{idx}"/><c:invertIfNegative val="0"/><c:bubble3D val="0"/>'
                f'<c:spPr><a:solidFill><a:srgbClr val="2D5576"/></a:solidFill><a:ln><a:noFill/></a:ln><a:effectLst/></c:spPr></c:dPt>'
            )
    s = s.replace('<c:dLbls>', dpts + '<c:dLbls>', 1)
    return s.encode('utf-8')

def _move_circle(slide_str, circle_name, new_x, new_y, new_cx, new_cy):
    idx = slide_str.find(f'name="{circle_name}"')
    if idx == -1: return slide_str
    start = slide_str.rfind('<p:pic>', 0, idx)
    end   = slide_str.find('</p:pic>', idx) + len('</p:pic>')
    pic   = slide_str[start:end]
    pic   = re.sub(r'<a:off x="[^"]*" y="[^"]*"/>', f'<a:off x="{new_x}" y="{new_y}"/>', pic)
    pic   = re.sub(r'<a:ext cx="[^"]*" cy="[^"]*"/>', f'<a:ext cx="{new_cx}" cy="{new_cy}"/>', pic)
    return slide_str[:start] + pic + slide_str[end:]

def _get_strat_circle_targets(strat_vals, threshold=0.3):
    """Pull/Push 각각 평균과 ±0.3 초과하는 것. 없으면 최대 1개. 합쳐서 3개 초과시 차이 큰 순으로 3개."""
    soft_avg = strat_vals[3]
    hard_avg = strat_vals[9]

    pull_diffs = [(i, abs(strat_vals[i] - soft_avg)) for i in range(3)]
    push_diffs = [(i, abs(strat_vals[i] - hard_avg)) for i in range(4, 9)]

    pull_over = [(i, d) for i, d in pull_diffs if d >= threshold]
    push_over = [(i, d) for i, d in push_diffs if d >= threshold]

    if not pull_over:
        pull_over = [max(pull_diffs, key=lambda x: x[1])]
    if not push_over:
        push_over = [max(push_diffs, key=lambda x: x[1])]

    combined = pull_over + push_over

    if len(combined) > 3:
        combined.sort(key=lambda x: -x[1])
        selected = combined[:3]
        # Pull/Push 각 최소 1개 보장
        if not any(i < 3 for i, _ in selected):
            best_pull = max(pull_over, key=lambda x: x[1])
            worst_push = min((x for x in selected if x[0] >= 4), key=lambda x: x[1])
            selected.remove(worst_push)
            selected.append(best_pull)
        if not any(i >= 4 for i, _ in selected):
            best_push = max(push_over, key=lambda x: x[1])
            worst_pull = min((x for x in selected if x[0] < 3), key=lambda x: x[1])
            selected.remove(worst_pull)
            selected.append(best_push)
        combined = selected

    return sorted([i for i, _ in combined])

def _update_circles(slide_str, comp_vals, strat_vals):
    # phase: circle2=최대값 막대, circle1=최소값 막대 (동값이면 각각 첫 번째)
    max_val = max(comp_vals); min_val = min(comp_vals)
    max_idx = comp_vals.index(max_val)
    min_idx = comp_vals.index(min_val)
    cw_p = int(_PHASE_BAR_W * 0.85)
    slide_str = _move_circle(slide_str, 'circle2',
        _bar_cx_phase(max_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)
    slide_str = _move_circle(slide_str, 'circle1',
        _bar_cx_phase(min_idx) - cw_p//2, _CIRCLE_PHASE_Y, cw_p, _CIRCLE_PHASE_CY)

    # strategy: Pull/Push 각 평균 기준 ±0.3, 최대 3개
    targets = _get_strat_circle_targets(strat_vals)
    cw_s = int(_STRAT_BAR_W * 0.85)
    for ci, data_idx in enumerate(targets):
        slide_str = _move_circle(slide_str, f'circle{ci+3}',
            _bar_cx_strat(data_idx) - cw_s//2, _CIRCLE_STRAT_Y, cw_s, _CIRCLE_STRAT_CY)
    return slide_str

def _fill_slide(sl_str, person, result):
    c=result["competency"]; s=result["skill_raw"]; sa=result["soft_avg"]; ha=result["hard_avg"]
    phase_data = list(c.items())
    strat_data = [(k,s[k]) for k in SOFT_SKILLS]+[("소프트스킬 평균",sa)]+[(k,s[k]) for k in HARD_SKILLS]+[("하드스킬 평균",ha)]
    comp_vals  = list(c.values())
    strat_vals = [s[k] for k in SOFT_SKILLS]+[sa]+[s[k] for k in HARD_SKILLS]+[ha]
    sl_str = sl_str.replace("{{NAME}}", person["name"])
    sl_str = _inject_table(sl_str, "table_phase",    phase_data)
    sl_str = _inject_table(sl_str, "table_strategy", strat_data)
    sl_str = _update_circles(sl_str, comp_vals, strat_vals)
    return sl_str

def build_ppt(people, ppt_tpl: bytes) -> bytes:
    with zipfile.ZipFile(io.BytesIO(ppt_tpl)) as src:
        infos = {info.filename: info for info in src.infolist()}
        files = {info.filename: src.read(info.filename) for info in src.infolist()}

    max_chart = max(int(m) for m in re.findall(r'chart(\d+)\.xml', ' '.join(files)))
    max_color = max(int(m) for m in re.findall(r'colors(\d+)\.xml', ' '.join(files)))
    max_style = max(int(m) for m in re.findall(r'style(\d+)\.xml', ' '.join(files)))
    ws_nums   = [int(m) if m else 0 for m in re.findall(r'Worksheet(\d*)\.xlsx', ' '.join(files))]
    max_ws    = max(ws_nums)

    prs_xml  = files["ppt/presentation.xml"]
    prs_rels = files["ppt/_rels/presentation.xml.rels"]
    ct_xml   = files["[Content_Types].xml"]
    max_sid  = max(int(m) for m in re.findall(r'<p:sldId id="(\d+)"', prs_xml.decode()))
    max_rid  = max(int(m) for m in re.findall(r'Id="rId(\d+)"', prs_rels.decode()))

    # slide2 기준 원본 (복제용)
    orig_s2  = files["ppt/slides/slide2.xml"]
    orig_s2r = files["ppt/slides/_rels/slide2.xml.rels"]
    orig_c3  = files["ppt/charts/chart3.xml"]
    orig_c4  = files["ppt/charts/chart4.xml"]
    orig_c3r = files["ppt/charts/_rels/chart3.xml.rels"]
    orig_c4r = files["ppt/charts/_rels/chart4.xml.rels"]

    for i, person in enumerate(people):
        result = compute(person["scores"])
        comp_vals  = list(result["competency"].values())
        strat_vals = ([result["skill_raw"][k] for k in SOFT_SKILLS] + [result["soft_avg"]] +
                      [result["skill_raw"][k] for k in HARD_SKILLS] + [result["hard_avg"]])

        if i == 0:
            sl = _fill_slide(files["ppt/slides/slide1.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide1.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart1.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart1.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart2.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart2.xml"], strat_vals), strat_vals)

        elif i == 1:
            sl = _fill_slide(files["ppt/slides/slide2.xml"].decode('utf-8'), person, result)
            files["ppt/slides/slide2.xml"] = sl.encode('utf-8')
            files["ppt/charts/chart3.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(files["ppt/charts/chart3.xml"], comp_vals), comp_vals)
            files["ppt/charts/chart4.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(files["ppt/charts/chart4.xml"], strat_vals), strat_vals)

        else:
            sn = i+1
            ca = max_chart+(i-1)*2+1; cb = ca+1
            cola = max_color+(i-1)*2+1; colb = cola+1
            sta  = max_style+(i-1)*2+1; stb  = sta+1
            wsa_n = max_ws+(i-1)*2+1;   wsb_n = wsa_n+1

            sl = _new_guids(orig_s2.decode('utf-8'))
            sl = _fill_slide(sl, person, result)
            files[f"ppt/slides/slide{sn}.xml"] = sl.encode('utf-8')
            files[f"ppt/slides/_rels/slide{sn}.xml.rels"] = (
                orig_s2r
                .replace(b"chart3.xml", f"chart{ca}.xml".encode())
                .replace(b"chart4.xml", f"chart{cb}.xml".encode())
            )
            files[f"ppt/charts/chart{ca}.xml"] = _update_chart_phase_colors(
                _replace_chart_vals(orig_c3, comp_vals), comp_vals)
            files[f"ppt/charts/chart{cb}.xml"] = _update_chart_strategy_colors(
                _replace_chart_vals(orig_c4, strat_vals), strat_vals)
            files[f"ppt/charts/_rels/chart{ca}.xml.rels"] = (
                orig_c3r
                .replace(b"chart3.xml",  f"chart{ca}.xml".encode())
                .replace(b"colors3.xml", f"colors{cola}.xml".encode())
                .replace(b"style3.xml",  f"style{sta}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet2.xlsx", _ws_name(wsa_n).encode())
            )
            files[f"ppt/charts/_rels/chart{cb}.xml.rels"] = (
                orig_c4r
                .replace(b"chart4.xml",  f"chart{cb}.xml".encode())
                .replace(b"colors4.xml", f"colors{colb}.xml".encode())
                .replace(b"style4.xml",  f"style{stb}.xml".encode())
                .replace(b"Microsoft_Excel_Worksheet3.xlsx", _ws_name(wsb_n).encode())
            )
            files[f"ppt/charts/colors{cola}.xml"] = files["ppt/charts/colors3.xml"]
            files[f"ppt/charts/colors{colb}.xml"] = files["ppt/charts/colors4.xml"]
            files[f"ppt/charts/style{sta}.xml"]   = files["ppt/charts/style3.xml"]
            files[f"ppt/charts/style{stb}.xml"]   = files["ppt/charts/style4.xml"]
            files[f"ppt/embeddings/{_ws_name(wsa_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"]
            files[f"ppt/embeddings/{_ws_name(wsb_n)}"] = files["ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"]

            def add_info(nn, rn):
                ni = zipfile.ZipInfo(nn); ni.compress_type = infos[rn].compress_type; infos[nn] = ni
            for nn, rn in [
                (f"ppt/slides/slide{sn}.xml",            "ppt/slides/slide2.xml"),
                (f"ppt/slides/_rels/slide{sn}.xml.rels", "ppt/slides/_rels/slide2.xml.rels"),
                (f"ppt/charts/chart{ca}.xml",            "ppt/charts/chart3.xml"),
                (f"ppt/charts/chart{cb}.xml",            "ppt/charts/chart4.xml"),
                (f"ppt/charts/_rels/chart{ca}.xml.rels", "ppt/charts/_rels/chart3.xml.rels"),
                (f"ppt/charts/_rels/chart{cb}.xml.rels", "ppt/charts/_rels/chart4.xml.rels"),
                (f"ppt/charts/colors{cola}.xml",         "ppt/charts/colors3.xml"),
                (f"ppt/charts/colors{colb}.xml",         "ppt/charts/colors4.xml"),
                (f"ppt/charts/style{sta}.xml",           "ppt/charts/style3.xml"),
                (f"ppt/charts/style{stb}.xml",           "ppt/charts/style4.xml"),
                (f"ppt/embeddings/{_ws_name(wsa_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet2.xlsx"),
                (f"ppt/embeddings/{_ws_name(wsb_n)}",    "ppt/embeddings/Microsoft_Excel_Worksheet3.xlsx"),
            ]: add_info(nn, rn)

            max_sid+=1; max_rid+=1; rid=f"rId{max_rid}"
            prs_xml  = prs_xml.replace(b'</p:sldIdLst>', f'<p:sldId id="{max_sid}" r:id="{rid}"/></p:sldIdLst>'.encode())
            prs_rels = prs_rels.replace(b'</Relationships>', f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{sn}.xml"/></Relationships>'.encode())
            ct_xml   = ct_xml.replace(b'</Types>',
                f'<Override PartName="/ppt/slides/slide{sn}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>'
                f'<Override PartName="/ppt/charts/chart{ca}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'<Override PartName="/ppt/charts/chart{cb}.xml" ContentType="application/vnd.openxmlformats-officedocument.drawingml.chart+xml"/>'
                f'</Types>'.encode())

    files["ppt/presentation.xml"]            = prs_xml
    files["ppt/_rels/presentation.xml.rels"] = prs_rels
    files["[Content_Types].xml"]             = ct_xml
    files["docProps/app.xml"] = re.sub(rb'<Slides>\d+</Slides>', f'<Slides>{len(people)}</Slides>'.encode(), files["docProps/app.xml"])

    out = io.BytesIO()
    with zipfile.ZipFile(out, 'w') as zout:
        for name, data in files.items():
            zout.writestr(infos[name], data)
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════
# 템플릿 탐색
# ══════════════════════════════════════════════════════════════════
def find_template(ext: str):
    base = Path(__file__).parent
    found = sorted(base.glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    found = sorted(Path(os.getcwd()).glob(f"*{ext}"))
    if found: return found[0].read_bytes(), str(found[0])
    return None, None

# ══════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="리더십 진단 보고서 자동화", layout="wide")
st.title("📊 리더십 진단 보고서 자동화 툴")

with st.sidebar:
    st.header("🔍 환경")
    try:
        files = [f.name for f in sorted(Path(__file__).parent.iterdir()) if f.is_file()]
        st.write("파일:", files)
    except Exception as e:
        st.write(f"오류: {e}")
    _, ep = find_template(".xlsx")
    _, pp = find_template(".pptx")
    st.write("📁 엑셀:", ep or "❌")
    st.write("📁 PPT:", pp or "❌")

st.markdown("---")
response_file = st.file_uploader("구글 폼 응답 엑셀 업로드", type=["xlsx","xls"])
st.markdown("---")

if st.button("🚀 보고서 생성", type="primary", use_container_width=True):
    if response_file is None:
        st.error("❌ 응답 엑셀을 업로드해주세요."); st.stop()

    resp_bytes = response_file.read()

    excel_tpl, ep = find_template(".xlsx")
    if not excel_tpl: st.error("❌ 엑셀 템플릿 없음 (GitHub 루트에 .xlsx 파일 필요)"); st.stop()

    ppt_tpl, pp = find_template(".pptx")
    if not ppt_tpl: st.error("❌ PPT 템플릿 없음 (GitHub 루트에 .pptx 파일 필요)"); st.stop()

    try:
        people = parse_people(resp_bytes)
    except Exception as e:
        st.error(f"❌ 파싱 실패: {e}"); st.code(traceback.format_exc()); st.stop()

    if not people: st.error("❌ 응답자 없음"); st.stop()

    st.info(f"👥 {len(people)}명: {[p['name'] for p in people]}")

    with st.expander("📋 점수 미리보기"):
        rows = []
        for p in people:
            r = compute(p["scores"])
            row = {"성함": p["name"]}
            row.update({k: f"{v:.2f}" for k,v in r["competency"].items()})
            row["소프트평균"] = r["soft_avg"]; row["하드평균"] = r["hard_avg"]
            rows.append(row)
        st.dataframe(pd.DataFrame(rows), use_container_width=True)

    with st.spinner(f"📊 엑셀 {len(people)}시트 생성 중..."):
        try:
            excel_out = build_excel(people, excel_tpl)
            st.success(f"✅ 엑셀 {len(people)}시트 완료")
        except Exception as e:
            st.error(f"❌ 엑셀 실패: {e}"); st.code(traceback.format_exc()); st.stop()

    with st.spinner(f"📑 PPT {len(people)}슬라이드 생성 중..."):
        try:
            ppt_out = build_ppt(people, ppt_tpl)
            st.success(f"✅ PPT {len(people)}슬라이드 완료")
        except Exception as e:
            st.error(f"❌ PPT 실패: {e}"); st.code(traceback.format_exc()); st.stop()

    st.session_state["excel_out"] = excel_out
    st.session_state["ppt_out"]   = ppt_out
    st.session_state["n"]         = len(people)
    st.session_state["done"]      = True

if st.session_state.get("done"):
    excel_out = st.session_state["excel_out"]
    ppt_out   = st.session_state["ppt_out"]
    n         = st.session_state["n"]

    st.success(f"🎉 완료: 엑셀 {n}시트 + PPT {n}슬라이드")

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("리더십진단_개인별.xlsx", excel_out)
        zf.writestr("리더십진단_통합.pptx",   ppt_out)

    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button("⬇️ ZIP (전체)", data=zip_buf.getvalue(),
            file_name="리더십진단_결과.zip", mime="application/zip", use_container_width=True)
    with d2:
        st.download_button("⬇️ 엑셀", data=excel_out,
            file_name="리더십진단_개인별.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
    with d3:
        st.download_button("⬇️ PPT", data=ppt_out,
            file_name="리더십진단_통합.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True)
